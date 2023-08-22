"""Crear un programa de banco que tenga la capacidad de crear una cuenta, hacer transferencia entre cuentas, hacer un deposito en la cuenta, 
hacer un retiro en la cuenta y consultar saldo en la cuenta"""

import openpyxl as openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from tkinter import *
import tkinter as tk

class programa:
    def __init__(self):
        self.ventana=Tk()
        self.title= "Banco de Calculo Numerico Enmanuel Díaz"
        self.size= "750x500"      

    def mostrar(self):
        self.ventana.mainloop()

    
    
        
    def cargar(self):
        ventana=self.ventana

        #tamaño de la ventana
        ventana.geometry(self.size)
        #titulo de la ventana
        ventana.title(self.title)
        #bienvenido a mi programa 
        texto= Label(ventana, text= "Bienvenido a nuestro Banco")
        texto.config(fg="White",
                bg="#000000",
                padx="750",
                pady="20",
                font=("Arial", 20))
        texto.pack()

        """Creación de 5 botones 1 para cada función"""

        #ventana y botón 1 de la creación de una cuenta bancaria
        
        def open_new_window():
            def añadir_datos():
                cedula = entry_cedula.get()
                nombre = entry_nombre.get()
                cantidad = entry_cantidad.get()

                wb = openpyxl.load_workbook('datos.xlsx')
                sheet = wb.active

                next_empty_row = sheet.max_row + 1
                sheet.cell(row=next_empty_row, column=1, value=cedula)
                sheet.cell(row=next_empty_row, column=2, value=nombre)
                sheet.cell(row=next_empty_row, column=3, value=cantidad)

                wb.save('datos.xlsx')
                wb.close()


            
            
            
            
            new_window = Toplevel(ventana)
            new_window.title("Crea una cuenta bancaria")
            new_window.geometry(self.size)
            texto= Label(new_window, text= "Crear Cuenta bancaria")
            texto.config(fg="White",
                bg="#000000",
                padx="750",
                pady="20",
                font=("Arial", 20))
            texto.pack()
            #entrada del nombre
            label_nombre = tk.Label(new_window, text="Nombre de la nueva cuenta:")
            label_nombre.pack(padx=10, pady=10)
            entry_nombre = Entry(new_window)
            entry_nombre.pack(padx=10, pady=5)

            #entrada de la cedula 
            label_nombre2 = tk.Label(new_window, text="cedula de la cuenta")
            label_nombre2.pack(padx=10, pady=10)
            entry_cedula = Entry(new_window)
            entry_cedula.pack(padx=10, pady=5)

            #entrada del dinero 
            label_nombre3 = tk.Label(new_window, text="Cantidad inicial de dinero de la cuenta")
            label_nombre3.pack(padx=10, pady=10)
            entry_cantidad = Entry(new_window)
            entry_cantidad.pack(padx=10, pady=5)

            #bóton para agregar valores
            botonsito = Button(new_window, text="Agregar Datos", command=añadir_datos)
            botonsito.pack(pady=10)
            #modulo para añadir datos 
            


           
             

            
            #new_label = ventana.Label(new_window, text="Creación de una cuenta bancaria")
            #new_label.pack(padx=20, pady=10)

        button = tk.Button(ventana, text="Crear cuenta bancaria", command=open_new_window)
        button.pack(padx=20, pady=10)
        #ventana y botón 2 del deposito normal
        def open_new_window2():

            def buscar_cedula():
                cedula_to_search = entry_cedula_buscar.get()

                wb = openpyxl.load_workbook('datos.xlsx')
                sheet = wb.active

                found_row_number = None
                for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    if cedula_to_search == row[0]:
                        found_row_number = row_number
                        break

                

                if found_row_number:
                    def sumar_deposito():
                        valor_de_deposito= entry_cedula_buscar_dime_dinero.get()
                        suma_de_deposito=float(sheet["C"+str(found_row_number)].value)+float(valor_de_deposito)
                        sheet["C"+str(found_row_number)]=round(suma_de_deposito, 2)
                        wb.save('datos.xlsx')
                    #result_label.config(text=f"Cédula encontrada en la fila: {found_row_number}")
                    
                    new_deposito = Toplevel(new_window)
                    new_deposito.title("Ventana del deposito")
                    
                    new_labeldeposito = Label(new_deposito, text="Indique cantidad de dinero a depositar en la cuenta")
                    new_labeldeposito.pack(padx=20, pady=10)

                    #preguntamos al cliente el dinero 
                    entry_cedula_buscar_dime_dinero = Entry(new_deposito)
                    entry_cedula_buscar_dime_dinero.pack(padx=10, pady=5)

                    button = Button(new_deposito, text="Hacer un deposito", command=sumar_deposito)
                    button.pack(padx=20, pady=10)
                else:
                    result_label.config(text="Cédula no encontrada")

                wb.close()
                
                            

                            

                    
                
            new_window = Toplevel(ventana)
            new_window.title("hacer un deposito")

            label_cedula = tk.Label(new_window, text="Diga la cedula de la cuenta a depositar")
            label_cedula.pack(padx=10, pady=10)
            entry_cedula_buscar = Entry(new_window)
            entry_cedula_buscar.pack(padx=10, pady=5)

            #el boton para buscar la cedula

            botonsito = Button(new_window, text="Hacer deposito", command=buscar_cedula)
            botonsito.pack(pady=10)
            result_label = Label(new_window, text="")
            result_label.pack()

        
            
            
            

        button = tk.Button(ventana, text="Hacer un deposito", command=open_new_window2)
        button.pack(padx=20, pady=10)
        
        
        #ventana y botón 3 del retiro de dinero, es muy parecido a deposito
        def open_new_window3():
            
            
            def buscar_cedula_retiro():
                cedula_to_search = entry_cedula_buscar.get()

                wb = openpyxl.load_workbook('datos.xlsx')
                sheet = wb.active

                found_row_number = None
                for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    if cedula_to_search == row[0]:
                        found_row_number = row_number
                        break

                

                if found_row_number:
                    def restar_retiro():
                        valor_de_retiro= entry_cedula_buscar_dime_dinero.get()
                        #la unica diferencia est[a en la resta
                        resta_del_retiro=float(sheet["C"+str(found_row_number)].value)-float(valor_de_retiro)
                        if resta_del_retiro>=0:
                            sheet["C"+str(found_row_number)]=round(resta_del_retiro, 2)
                            wb.save('datos.xlsx')
                        else:
                            new_labelretiro_respuestanegativa = Label(new_deposito, text="Usted no tiene dinero suficiente en la cuenta")
                            new_labelretiro_respuestanegativa.pack(padx=20, pady=10)

                    #result_label.config(text=f"Cédula encontrada en la fila: {found_row_number}")
                    
                    new_deposito = Toplevel(new_window)
                    new_deposito.title("Ventana del deposito")
                    
                    new_labeldeposito = Label(new_deposito, text="Indique cantidad de dinero a retirar en la cuenta")
                    new_labeldeposito.pack(padx=20, pady=10)

                    #preguntamos al cliente el dinero 
                    entry_cedula_buscar_dime_dinero = Entry(new_deposito)
                    entry_cedula_buscar_dime_dinero.pack(padx=10, pady=5)

                    button = Button(new_deposito, text="Retiro", command=restar_retiro)
                    button.pack(padx=20, pady=10)
                else:
                    result_label.config(text="Cédula no encontrada")

                wb.close()
                
                            

                            

                    
                
            new_window = Toplevel(ventana)
            new_window.title("hacer un retiro")

            label_cedula = tk.Label(new_window, text="Diga la cedula de la cuenta a retirar")
            label_cedula.pack(padx=10, pady=10)
            entry_cedula_buscar = Entry(new_window)
            entry_cedula_buscar.pack(padx=10, pady=5)

            #el boton para buscar la cedula

            botonsito = Button(new_window, text="Hacer retiro", command=buscar_cedula_retiro)
            botonsito.pack(pady=10)
            result_label = Label(new_window, text="")
            result_label.pack()

        
            
            
            

        button = tk.Button(ventana, text="Hacer un retiro", command=open_new_window3)
        button.pack(padx=20, pady=10)











        #Ventana y botón 4 de la trasnferencia de dinero
        def open_new_window4():

            #consulta de las cedulas y proceso despues de boton 
            def buscar_cedula_del_que_envia():
                cedula_to_search = entre_cedula_de_hace_trasnferencia.get()

                wb = openpyxl.load_workbook('datos.xlsx')
                sheet = wb.active

                found_row_number = None
                for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    if cedula_to_search == row[0]:
                        found_row_number = row_number
                        break

                if found_row_number:
                    
                        valor_de_transferencia= entre_monto_envio.get()
                        #la unica diferencia esta en la resta
                        resta_de_la_transferencia =float(sheet["C"+str(found_row_number)].value)-float(valor_de_transferencia)
                        if resta_de_la_transferencia>=0:
                            sheet["C"+str(found_row_number)]=round(resta_de_la_transferencia, 2)
                            wb.save('datos.xlsx')

                            cedula_to_search = entre_cedula_de_recibe_trasnferencia.get()
                            found_row_number = None
                            for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                                if cedula_to_search == row[0]:
                                    found_row_number = row_number
                                    break

                            suma_de_transferencia= float(valor_de_transferencia) + float(sheet["C"+str(found_row_number)].value)
                            sheet["C"+str(found_row_number)]=round(suma_de_transferencia, 2)
                            wb.save('datos.xlsx')
                        else:
                            new_labeltrasnferencia_respuestanegativa = Label(new_window, text="Usted no tiene dinero suficiente en la cuenta")
                            new_labeltrasnferencia_respuestanegativa.pack(padx=20, pady=10)

                    
                    
                    
                else:
                    result_label=Label(new_window, text="Cedula no encontrada")
                    result_label.pack(padx=20, pady=10)



            #separacion
            new_window = Toplevel(ventana)
            new_window.title("Nueva Ventana4")

            #pido informacion del que hace la transferencia
            
            new_label = Label(new_window, text="Diga cedula de quien envia")
            new_label.pack(padx=20, pady=10)
            entre_cedula_de_hace_trasnferencia= Entry(new_window)
            entre_cedula_de_hace_trasnferencia.pack(padx=20, pady=10)

            #pido informacion del que recibe

            new_label2 = Label(new_window, text="Diga cedula de quien recibe")
            new_label2.pack(padx=20, pady=10)
            entre_cedula_de_recibe_trasnferencia= Entry(new_window)
            entre_cedula_de_recibe_trasnferencia.pack(padx=20, pady=10)

            #valor del dinero que se va ha enviar
            new_label3 = Label(new_window, text="Diga el monto del envio")
            new_label3.pack(padx=20, pady=10)
            entre_monto_envio= Entry(new_window)
            entre_monto_envio.pack(padx=20, pady=10)

            el_boton= tk.Button(new_window, text="procesar envio",command=buscar_cedula_del_que_envia )
            el_boton.pack(padx=20, pady=10)

            #validamos el dinero del que hace la transferencia

            



        button = tk.Button(ventana, text="Hacer una transferencia", command=open_new_window4)
        button.pack(padx=20, pady=10)





        #ventana 5 de la consulta de dinero
        def open_new_window5():
            new_window = Toplevel(ventana)
            new_window.title("Ventana de consulta de saldo")
            def buscar_cedula_consulta():
                cedula_to_search = entry_cedula_consulta.get()

                wb = openpyxl.load_workbook('datos.xlsx')
                sheet = wb.active

                found_row_number = None
                for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    if cedula_to_search == row[0]:
                        found_row_number = row_number
                        break
                if found_row_number:
                    consulta= "usted tiene en la cuenta "+ str(sheet["C"+str(found_row_number)].value)
                    label_de_la_consulta= Label(new_window,text=consulta)
                    label_de_la_consulta.pack(padx=20, pady=10)
                else:
                    label_de_la_consulta= Label(new_window, text="usted no tiene dinero")
                    label_de_la_consulta.pack(padx=20, pady=10)
                    
            
            new_label = Label(new_window, text="diga la cedula de la consulta")
            new_label.pack(padx=20, pady=10)

            entry_cedula_consulta = Entry(new_window)
            entry_cedula_consulta.pack(padx=10, pady=5)

            botonsito = Button(new_window, text="Hacer consulta", command=buscar_cedula_consulta)
            botonsito.pack(pady=10)

        button = tk.Button(ventana, text="Hacer una consulta de saldo", command=open_new_window5)
        button.pack(padx=20, pady=10)


        
programa= programa()
programa.cargar()
programa.mostrar()
            








